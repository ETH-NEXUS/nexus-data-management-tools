import csv
from os.path import isabs, join
from typing import Any, Dict, List, Tuple, Optional

# Package-safe imports with fallback for script execution
try:  # LabKey and helpers
    from .helpers import Message as M
    from labkey.api_wrapper import APIWrapper
    from labkey.query import QueryFilter
    from labkey.exceptions import (
        RequestError,
        QueryNotFoundError,
        ServerContextError,
        ServerNotFoundError,
    )
except ImportError:  # pragma: no cover
    from helpers import Message as M  # type: ignore
    from labkey.api_wrapper import APIWrapper  # type: ignore
    from labkey.query import QueryFilter  # type: ignore
    from labkey.exceptions import (  # type: ignore
        RequestError,
        QueryNotFoundError,
        ServerContextError,
        ServerNotFoundError,
    )

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency validated at runtime
    load_workbook = None  # type: ignore


def _resolve_path(drop_folder: str, path: str) -> str:
    return path if isabs(path) else join(drop_folder, path)


def _load_labkey_rows(cfg: Dict[str, Any]) -> List[Dict[str, Any]]:
    api = APIWrapper(cfg["host"], cfg["container"], use_ssl=True)
    filters = []
    for f in cfg.get("filters", []) or []:
        ftype = (f.get("type") or "").lower()
        fval = f.get("value")
        fname = f.get("field")
        if not (fname and fval is not None):
            continue
        if ftype == "contains":
            filters.append(QueryFilter(fname, fval, QueryFilter.Types.CONTAINS))
        elif ftype in ("eq", "equals", "="):
            filters.append(QueryFilter(fname, fval, QueryFilter.Types.EQUAL))
        else:
            # default to EQUAL
            filters.append(QueryFilter(fname, fval, QueryFilter.Types.EQUAL))
    result = api.query.select_rows(
        cfg["schema"], cfg["table"], columns=cfg.get("columns"), filter_array=filters
    )
    return result.get("rows", [])


def _load_excel_rows(path: str, sheet: Optional[str] = None) -> List[Dict[str, Any]]:
    if load_workbook is None:
        raise RuntimeError(
            "openpyxl is required to read Excel files. Please install and configure it."
        )
    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb.active
    rows: List[Dict[str, Any]] = []
    header = [str(v) if v is not None else "" for v in next(ws.iter_rows(values_only=True))]
    for r in ws.iter_rows(min_row=2, values_only=True):
        row_dict = {header[i]: r[i] for i in range(len(header))}
        rows.append(row_dict)
    return rows


def _load_csv_rows(path: str, delimiter: str = ",", encoding: str = "utf-8") -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    with open(path, "r", encoding=encoding, newline="") as fh:
        reader = csv.DictReader(fh, delimiter=delimiter)
        for row in reader:
            rows.append(dict(row))
    return rows


def load_metadata_sources(sources_cfg: List[Dict[str, Any]], drop_folder: str) -> List[Dict[str, Any]]:
    """Load metadata rows for each configured source.

    sources_cfg: list of sources, each with a 'type' and type-specific fields.
    Returns a list of dicts: { name, type, count, status, rows?, error? }
    """
    results: List[Dict[str, Any]] = []

    for src in sources_cfg or []:
        stype = (src.get("type") or "").lower()
        name = src.get("name") or stype
        status = "ok"
        error: Optional[str] = None
        rows: List[Dict[str, Any]] = []
        try:
            if stype == "labkey":
                rows = _load_labkey_rows(src)
            elif stype == "excel":
                path = _resolve_path(drop_folder, src["path"])
                rows = _load_excel_rows(path, sheet=src.get("sheet"))
            elif stype == "csv":
                path = _resolve_path(drop_folder, src["path"])
                rows = _load_csv_rows(path, delimiter=src.get("delimiter", ","))
            else:
                status = "error"
                error = f"Unknown source type: {stype}"
        except (ServerContextError, ServerNotFoundError, QueryNotFoundError, RequestError) as ex:
            status = "error"
            error = f"LabKey error: {ex}"
        except Exception as ex:  # pylint: disable=broad-except
            status = "error"
            error = str(ex)
        if status == "error":
            M.warn(f"Metadata source '{name}' failed: {error}")
        results.append(
            {
                "name": name,
                "type": stype,
                "count": len(rows),
                "status": status,
                "error": error,
                "rows": rows,
            }
        )

    return results
